import json
import os
import sys
import tempfile
from pathlib import Path

from django.db import IntegrityError
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from ..models import Category, Product
from .helpers import _build_xlsx_response
from .import_helpers import (
    IMPORT_COLUMNS,
    _detect_duplicate_barcodes,
    _extract_and_persist,
    _match_product,
    _validate_row,
)


# ---------------------------------------------------------------------------
# Template View  (GET  /api/import/products/template/)
# ---------------------------------------------------------------------------

class ImportProductsTemplateView(APIView):
    """Download an XLSX template with **default products** from the fixture
    so the user can edit prices, costs, stock, etc. before uploading."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        headers = [
            "ID",
            "Nombre",
            "Categoría",
            "Precio Venta",
            "Costo",
            "Stock",
            "Stock Mínimo",
            "Código Barras",
            "Descripción",
        ]
        ws.append(headers)

        # ── Load fixture products ──────────────────────────────────────
        # Resolve fixture path (same logic as load_initial_data command)
        if getattr(sys, 'frozen', False):
            fixture_path = Path(sys._MEIPASS) / "coreApp" / "fixtures" / "initial_data.json"
        else:
            # imports.py → coreApp/views/ → coreApp/ → coreApp/fixtures/initial_data.json
            fixture_path = (
                Path(__file__).parent.parent
                / "fixtures"
                / "initial_data.json"
            )

        if fixture_path.exists():
            with open(fixture_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            # Build category PK → name lookup
            fixture_cats = {
                item["pk"]: item["fields"]["name"]
                for item in data
                if item["model"] == "coreApp.category"
            }

            # Append each fixture product as a row
            for item in data:
                if item["model"] != "coreApp.product":
                    continue
                fields = item["fields"]
                cat_name = fixture_cats.get(fields["category"], "")
                ws.append([
                    "",                              # ID (vacío — se creará nuevo)
                    fields["name"],
                    cat_name,
                    fields["price"],
                    fields.get("cost", "0.00"),
                    fields.get("stock", 0),
                    fields.get("min_stock", 10),
                    fields.get("barcode", ""),
                    fields.get("description", ""),
                ])

        return _build_xlsx_response(wb, "plantilla-productos.xlsx")


# ---------------------------------------------------------------------------
# Import View  (POST  /api/import/products/)
# ---------------------------------------------------------------------------

class ImportProductsView(APIView):
    """Accept an ``.xlsx`` file, match/create products, and report results.

    Pass ``?preview=true`` to validate without persisting any changes.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        preview = request.query_params.get("preview", "").lower() == "true"
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"detail": "No se envió ningún archivo"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not file.name.endswith(".xlsx"):
            return Response(
                {"detail": "Formato inválido — se espera un archivo .xlsx"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Write upload to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            wb = load_workbook(tmp_path)
            ws = wb.active

            # ---- Read headers → build column map ----
            headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip() if h is not None else "" for h in headers_row]
            col_map = {h: i for i, h in enumerate(headers) if h}

            # Validate required columns are present
            missing = [c for c in IMPORT_COLUMNS["required"] if c not in col_map]
            if missing:
                return Response(
                    {"detail": f"Faltan columnas requeridas: {', '.join(missing)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Collect data rows (row 2 onward)
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))

            # Reject completely empty file
            has_data = any(any(cell is not None for cell in r) for r in all_rows)
            if not has_data:
                return Response(
                    {"detail": "El archivo no contiene datos"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # ---- Pre-scan barcode duplicates ----
            duplicate_indices = _detect_duplicate_barcodes(all_rows, col_map)

            created = 0
            updated = 0
            errors: list[dict] = []

            # ---- Process rows ----
            for i, row in enumerate(all_rows):
                if all(cell is None for cell in row):
                    continue

                excel_row = i + 2
                row_errors: list[str] = []

                if i in duplicate_indices:
                    row_errors.append("Código Barras duplicado en el archivo")

                row_errors.extend(_validate_row(row, col_map))

                if row_errors:
                    for msg in row_errors:
                        errors.append({"row": excel_row, "message": msg})
                    continue

                try:
                    product = _match_product(row, col_map)
                except ValueError as exc:
                    errors.append({"row": excel_row, "message": str(exc)})
                    continue

                result = _extract_and_persist(row, col_map, product, preview)
                if result["error"]:
                    errors.append({"row": excel_row, "message": result["error"]})
                elif result["action"] == "created":
                    created += 1
                elif result["action"] == "updated":
                    updated += 1

            result = {"created": created, "updated": updated, "errors": errors}
            if preview:
                result["preview"] = True
            return Response(result)

        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)