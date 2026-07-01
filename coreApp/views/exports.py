from rest_framework.views import APIView
from rest_framework.response import Response

from ..models import Client, Expense, Product, Sale
from .helpers import _build_xlsx_response


class ExportClientsView(APIView):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        headers = ["ID", "Nombre", "Teléfono", "Email", "Dirección", "Límite Crédito", "Deuda Actual"]
        ws.append(headers)

        for client in Client.objects.all().order_by("name").iterator():
            ws.append(
                [
                    client.id,
                    client.name,
                    client.phone or "",
                    client.email or "",
                    client.address or "",
                    str(client.credit_limit),
                    str(client.current_debt),
                ]
            )

        return _build_xlsx_response(wb, "clientes.xlsx")


class ExportProductsView(APIView):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        headers = [
            "ID",
            "Nombre",
            "Categoría",
            "Precio Venta",
            "Costo",
            "Stock",
            "Stock Mínimo",
            "Código Barras",
        ]
        ws.append(headers)

        for product in Product.objects.all().order_by("name").select_related("category").iterator():
            ws.append(
                [
                    product.id,
                    product.name,
                    product.category.name if product.category else "",
                    str(product.price),
                    str(product.cost),
                    product.stock,
                    product.min_stock,
                    product.barcode or "",
                ]
            )

        return _build_xlsx_response(wb, "productos.xlsx")


class ExportSalesView(APIView):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        headers = [
            "ID Venta",
            "Fecha",
            "Cliente",
            "Total",
            "Método Pago",
            "Estado",
            "Notas",
            "Efectivo Recibido",
            "Vuelto",
        ]
        ws.append(headers)

        qs = Sale.objects.all().order_by("-created_at").select_related("client")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        for sale in qs.iterator():
            ws.append(
                [
                    sale.id,
                    sale.created_at.strftime("%Y-%m-%d %H:%M"),
                    sale.client.name if sale.client else "—",
                    str(sale.total),
                    sale.get_payment_method_display(),
                    sale.get_status_display(),
                    sale.notes or "",
                    str(sale.cash_received or ""),
                    str(sale.change_given or ""),
                ]
            )

        return _build_xlsx_response(wb, "ventas.xlsx")


class ExportExpensesView(APIView):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"

        headers = ["ID", "Fecha", "Categoría", "Descripción", "Monto", "Creado"]
        ws.append(headers)

        qs = Expense.objects.all().order_by("-date", "-created_at")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        for expense in qs.iterator():
            ws.append(
                [
                    expense.id,
                    expense.date.strftime("%Y-%m-%d"),
                    expense.get_category_display(),
                    expense.description,
                    str(expense.amount),
                    expense.created_at.strftime("%Y-%m-%d %H:%M"),
                ]
            )

        return _build_xlsx_response(wb, "gastos.xlsx")