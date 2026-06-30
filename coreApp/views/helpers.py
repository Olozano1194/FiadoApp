from datetime import datetime, timedelta
from decimal import Decimal

from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone

from ..models import Expense, FiadoPayment, Sale, SaleItem


def get_week_range(date_str: str | None = None) -> tuple[datetime, datetime]:
    """Return (week_start, week_end) for the given iso‑date string or current week."""
    if date_str:
        ref = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        ref = timezone.localdate()
    start = ref - timedelta(days=ref.weekday())
    end = start + timedelta(days=7)
    return (
        timezone.make_aware(datetime.combine(start, datetime.min.time())),
        timezone.make_aware(datetime.combine(end, datetime.min.time())),
    )


def _build_xlsx_response(workbook, filename: str) -> HttpResponse:
    """Wrap an openpyxl workbook in an HttpResponse with proper headers."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # Style the header row
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003527", end_color="003527", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws = workbook.active
    ws.title = filename.replace(".xlsx", "")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Auto‑width columns
    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = col_cells[0].column_letter
        lengths = []
        for cell in col_cells:
            try:
                lengths.append(len(str(cell.value)))
            except Exception:
                lengths.append(0)
        ws.column_dimensions[col_letter].width = min(max(lengths or [10]) + 2, 40)

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def calculate_closure_data(date):
    """Calculate closure financial data for a given date.

    Args:
        date: A ``datetime.date`` (or ``datetime.datetime``) for the target day.

    Returns:
        dict with the following Decimal/int keys:
            cash_sales, credit_sales, total_sales, sales_count,
            fiado_payments, expenses, net_profit, expected_cash
    """
    today_start = timezone.make_aware(datetime.combine(date, datetime.min.time()))
    today_end = today_start + timedelta(days=1)

    cash_sales = (
        Sale.objects.filter(
            created_at__range=(today_start, today_end),
            status="COMPLETED",
            payment_method="CASH",
        ).aggregate(total=Sum("total"))["total"]
        or Decimal("0.00")
    )

    credit_sales = (
        Sale.objects.filter(
            created_at__range=(today_start, today_end),
            status="COMPLETED",
            payment_method="CREDIT",
        ).aggregate(total=Sum("total"))["total"]
        or Decimal("0.00")
    )

    total_sales = cash_sales + credit_sales

    sales_count = Sale.objects.filter(
        created_at__range=(today_start, today_end), status="COMPLETED"
    ).count()

    fiado_payments = (
        FiadoPayment.objects.filter(date__range=(today_start, today_end)).aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )

    expenses = (
        Expense.objects.filter(date=date).aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )

    today_sales = Sale.objects.filter(
        created_at__range=(today_start, today_end), status="COMPLETED"
    )

    net_profit = Decimal("0.00")
    for sale in today_sales:
        for item in sale.items.all():
            cost = item.cost_at_sale or Decimal("0")
            net_profit += (item.unit_price - cost) * item.quantity

    expected_cash = cash_sales + fiado_payments - expenses

    return {
        "cash_sales": cash_sales,
        "credit_sales": credit_sales,
        "total_sales": total_sales,
        "sales_count": sales_count,
        "fiado_payments": fiado_payments,
        "expenses": expenses,
        "net_profit": net_profit,
        "expected_cash": expected_cash,
    }
