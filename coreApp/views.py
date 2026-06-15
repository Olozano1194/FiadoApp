import os
from datetime import datetime, timedelta
from decimal import Decimal
from itertools import chain
from collections import defaultdict

from django.conf import settings
from django.db.models import F, Q, Sum, Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .backup_utils import backup_db, detect_engine, get_db_file_size, restore_db
from .models import (
    BackupConfig,
    CashClosure,
    Category,
    Client,
    Expense,
    FiadoPayment,
    Product,
    Sale,
    SaleItem,
)
from .serializers import (
    CashClosureCreateSerializer,
    CashClosureSerializer,
    CategorySerializer,
    ClientSerializer,
    ExpenseSerializer,
    FiadoPaymentSerializer,
    ProductSerializer,
    SaleCreateSerializer,
    SaleItemSerializer,
    SaleSerializer,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def get_week_range(date_str: str | None = None) -> tuple[datetime, datetime]:
    """Return (week_start, week_end) for the given iso‑date string or current week."""
    if date_str:
        ref = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        ref = timezone.localdate()
    start = ref - timedelta(days=ref.weekday())
    end = start + timedelta(days=7)
    return (
        timezone.make_aware(datetime.combine(start, datetime.min.time())),
        timezone.make_aware(datetime.combine(end, datetime.min.time())),
    )


def _build_xlsx_response(workbook, filename: str) -> HttpResponse:
    """Wrap an openpyxl workbook in an HttpResponse with proper headers."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # Style the header row
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003527", end_color="003527", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws = workbook.active
    ws.title = filename.replace(".xlsx", "")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Auto‑width columns
    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = col_cells[0].column_letter
        lengths = []
        for cell in col_cells:
            try:
                lengths.append(len(str(cell.value)))
            except Exception:
                lengths.append(0)
        ws.column_dimensions[col_letter].width = min(max(lengths or [10]) + 2, 40)

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


# ---------------------------------------------------------------------------
# Standard  Pagination
# ---------------------------------------------------------------------------


class StandardPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = "page_size"
    max_page_size = 100


# ---------------------------------------------------------------------------
# Health Check
# ---------------------------------------------------------------------------


class HealthCheckView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        return Response({"status": "ok", "message": "FiadoApp API running"})


# ---------------------------------------------------------------------------
# Category
# ---------------------------------------------------------------------------


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all().order_by("name")
    serializer_class = CategorySerializer
    pagination_class = None


# ---------------------------------------------------------------------------
# Product
# ---------------------------------------------------------------------------


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all().order_by("name")
    serializer_class = ProductSerializer

    @action(detail=False, url_path="low-stock")
    def low_stock(self, request):
        """Return products where stock is below min_stock."""
        products = Product.objects.filter(stock__lt=F("min_stock")).order_by("name")
        page = self.paginate_queryset(products)
        if page is not None:
            serializer = ProductSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = ProductSerializer(products, many=True)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# Client
# ---------------------------------------------------------------------------


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all().order_by("name")
    serializer_class = ClientSerializer

    def get_queryset(self):
        qs = Client.objects.all().order_by("name")
        q = self.request.query_params.get("q")
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(phone__icontains=q) | Q(email__icontains=q))
        debt = self.request.query_params.get("debt")
        if debt == "true":
            qs = qs.filter(current_debt__gt=0)
        return qs


# ---------------------------------------------------------------------------
# Sale
# ---------------------------------------------------------------------------


class SaleViewSet(viewsets.ModelViewSet):
    queryset = Sale.objects.all().order_by("-created_at")
    serializer_class = SaleSerializer
    pagination_class = StandardPagination

    def get_serializer_class(self):
        if self.action in ("create", "update", "partial_update"):
            return SaleCreateSerializer
        return SaleSerializer

    @action(detail=False, methods=['get'], url_path='recent')
    def recent(self, request):
        """Return recent completed sales."""
        limit = int(request.query_params.get("limit", 10))
        qs = (
            Sale.objects.filter(status="COMPLETED")
            .order_by("-created_at")
            .select_related("client")[:limit]
        )
        data = [
            {
                "id": sale.id,
                "cliente": sale.client.name if sale.client else "—",
                "hora": sale.created_at.strftime("%H:%M"),
                "estado": sale.get_status_display(),
                "total": str(sale.total),
            }
            for sale in qs
        ]
        return Response(data)

    @action(detail=False)
    def history(self, request):
        """Return sales. If ?client_id=N, filter by client + COMPLETED. Otherwise ALL sales."""
        client_id = request.query_params.get("client_id")
        if client_id:
            qs = Sale.objects.filter(client_id=client_id, status="COMPLETED").select_related("client").order_by("-created_at")
        else:
            qs = Sale.objects.all().order_by("-created_at").select_related("client")
        page = self.paginate_queryset(qs)
        if page is not None:
            results = [
                {
                    "id": sale.id,
                    "cliente": sale.client.name if sale.client else "—",
                    "fecha": sale.created_at.strftime("%Y-%m-%d"),
                    "hora": sale.created_at.strftime("%H:%M"),
                    "metodo_pago": sale.get_payment_method_display(),
                    "estado": sale.get_status_display(),
                    "total": str(sale.total),
                }
                for sale in page
            ]
            return self.get_paginated_response(results)
        return Response({"detail": "No page?"}, status=400)


# ---------------------------------------------------------------------------
# Fiado Payment
# ---------------------------------------------------------------------------


class FiadoPaymentViewSet(viewsets.ModelViewSet):
    queryset = FiadoPayment.objects.all().order_by("-date")
    serializer_class = FiadoPaymentSerializer

    def perform_create(self, serializer):
        payment = serializer.save()
        # Update client's current_debt
        if payment.client:
            payment.client.current_debt -= payment.amount
            payment.client.save(update_fields=["current_debt"])

    @action(detail=False)
    def today(self, request):
        """Return today's payment summary (total and count)."""
        today_start = timezone.make_aware(
            datetime.combine(timezone.localdate(), datetime.min.time())
        )
        today_end = today_start + timedelta(days=1)
        qs = FiadoPayment.objects.filter(date__range=(today_start, today_end))
        total = qs.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        count = qs.count()
        return Response({"total": str(total), "count": count})


# ---------------------------------------------------------------------------
# Expense
# ---------------------------------------------------------------------------


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all().order_by("-date", "-created_at")
    serializer_class = ExpenseSerializer
    pagination_class = None


# ---------------------------------------------------------------------------
# Cash Closure
# ---------------------------------------------------------------------------


class CashClosureViewSet(viewsets.ModelViewSet):
    queryset = CashClosure.objects.all().order_by("-date")
    serializer_class = CashClosureSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == "create":
            return CashClosureCreateSerializer
        if self.action == "preview":
            return CashClosureCreateSerializer
        return CashClosureSerializer

    def create(self, request, *args, **kwargs):
        """CashClosureCreateSerializer for input, CashClosureSerializer for output."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save(created_by=request.user)
        out = CashClosureSerializer(instance, context=self.get_serializer_context())
        return Response(out.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"])
    def preview(self, request):
        """Return a preview of today's closure without persisting it."""
        today = timezone.localdate()
        today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        today_end = today_start + timedelta(days=1)

        cash_sales = (
            Sale.objects.filter(
                created_at__range=(today_start, today_end),
                status="COMPLETED",
                payment_method="CASH",
            ).aggregate(total=Sum("total"))["total"]
            or Decimal("0.00")
        )

        credit_sales = (
            Sale.objects.filter(
                created_at__range=(today_start, today_end),
                status="COMPLETED",
                payment_method="CREDIT",
            ).aggregate(total=Sum("total"))["total"]
            or Decimal("0.00")
        )

        total_sales = cash_sales + credit_sales

        sales_count = Sale.objects.filter(
            created_at__range=(today_start, today_end), status="COMPLETED"
        ).count()

        fiado_payments = (
            FiadoPayment.objects.filter(date__range=(today_start, today_end)).aggregate(
                total=Sum("amount")
            )["total"]
            or Decimal("0.00")
        )

        expenses = (
            Expense.objects.filter(date=today).aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00")
        )

        today_sales = Sale.objects.filter(
            created_at__range=(today_start, today_end), status="COMPLETED"
        ).prefetch_related("items__product")

        net_profit = Decimal("0.00")
        for sale in today_sales:
            for item in sale.items.all():
                cost = item.product.cost or Decimal("0")
                net_profit += (item.unit_price - cost) * item.quantity

        expected_cash = cash_sales + fiado_payments - expenses

        data = {
            "date": today.isoformat(),
            "total_sales": str(total_sales),
            "cash_sales": str(cash_sales),
            "credit_sales": str(credit_sales),
            "sales_count": sales_count,
            "fiado_payments": str(fiado_payments),
            "expenses": str(expenses),
            "net_profit": str(net_profit),
            "expected_cash": str(expected_cash),
        }

        # Check if a closure already exists for today
        try:
            existing = CashClosure.objects.get(date=today)
            data["already_closed"] = True
            data["last_closure"] = existing.created_at.isoformat()
            data["last_counted_cash"] = str(existing.counted_cash)
            data["last_discrepancy"] = str(existing.discrepancy)
        except CashClosure.DoesNotExist:
            data["already_closed"] = False

        return Response(data)


# ---------------------------------------------------------------------------
# Dashboard Stats
# ---------------------------------------------------------------------------


class DashboardStatsView(APIView):
    def get(self, request):
        today = timezone.localdate()
        today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        today_end = today_start + timedelta(days=1)

        # Today's sales
        today_sales_qs = Sale.objects.filter(
            created_at__range=(today_start, today_end), status="COMPLETED"
        )
        today_total = today_sales_qs.aggregate(total=Sum("total"))["total"] or Decimal("0.00")

        # Today's expenses
        today_expenses = (
            Expense.objects.filter(date=today).aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00")
        )

        # Yesterday's sales for comparison
        yesterday_start = today_start - timedelta(days=1)
        yesterday_end = today_start
        yesterday_total = (
            Sale.objects.filter(
                created_at__range=(yesterday_start, yesterday_end), status="COMPLETED"
            ).aggregate(total=Sum("total"))["total"]
            or Decimal("0.00")
        )

        # cambio_vs_ayer
        if yesterday_total > 0:
            cambio_pct = float(((today_total - yesterday_total) / yesterday_total) * 100)
            cambio_vs_ayer = f"{cambio_pct:+.1f}"
        elif today_total > 0:
            cambio_vs_ayer = "+100.0"
        else:
            cambio_vs_ayer = "0.0"

        # Total clients with debt
        clients_with_debt = Client.objects.filter(current_debt__gt=0).count()
        total_debt = Client.objects.aggregate(total=Sum("current_debt"))["total"] or Decimal("0.00")

        # Low stock count
        low_stock_count = Product.objects.filter(stock__lt=F("min_stock")).count()

        # Net today / ganancia_dia
        net_today = today_total - today_expenses

        # margen_dia
        if today_total > 0:
            margen_dia = round(float((net_today / today_total) * 100), 1)
        else:
            margen_dia = None

        # Weekly sales trend (last 7 days)
        week_ago = today_start - timedelta(days=7)
        weekly_sales = []
        for i in range(7):
            day = today_start - timedelta(days=i)
            day_end = day + timedelta(days=1)
            day_total = (
                Sale.objects.filter(
                    created_at__range=(day, day_end), status="COMPLETED"
                ).aggregate(total=Sum("total"))["total"]
                or Decimal("0.00")
            )
            weekly_sales.append({"date": day.date().isoformat(), "total": str(day_total)})

        return Response(
            {
                "ventas_dia": str(today_total),
                "gastos_hoy": str(today_expenses),
                "cambio_vs_ayer": cambio_vs_ayer,
                "ganancia_dia": str(net_today),
                "margen_dia": margen_dia,
                "fiado_pendiente_total": str(total_debt),
                "clientes_fiado_pendiente": clients_with_debt,
                "productos_stock_bajo": low_stock_count,
                "weekly_sales": weekly_sales,
            }
        )


# ---------------------------------------------------------------------------
# Search (global)
# ---------------------------------------------------------------------------


class SearchView(APIView):
    def get(self, request):
        q = request.query_params.get("q", "")
        if not q or len(q) < 2:
            return Response({"products": [], "clients": [], "sales": []})

        products = Product.objects.filter(
            Q(name__icontains=q) | Q(barcode__icontains=q)
        )[:10]
        clients = Client.objects.filter(name__icontains=q)[:10]
        sales = Sale.objects.filter(
            Q(client__name__icontains=q) | Q(id__icontains=q)
        ).select_related("client").prefetch_related("items__product")[:10]

        return Response(
            {
                "products": ProductSerializer(products, many=True).data,
                "clients": ClientSerializer(clients, many=True).data,
                "sales": SaleSerializer(sales, many=True).data,
            }
        )


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------


class ReportStatsView(APIView):
    def get(self, request):
        today = timezone.localdate()
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        week_str = request.query_params.get("week")

        if week_str:
            week_start, week_end = get_week_range(week_str)
            range_start, range_end = week_start, week_end
            report_start = week_start.date()
            report_end = (week_end - timedelta(days=1)).date()
        elif date_from and date_to:
            range_end = timezone.make_aware(
                datetime.combine(datetime.strptime(date_to, "%Y-%m-%d").date(), datetime.max.time())
            )
            range_start = timezone.make_aware(
                datetime.combine(
                    datetime.strptime(date_from, "%Y-%m-%d").date(), datetime.min.time()
                )
            )
            report_start = range_start.date()
            report_end = range_end.date()
        else:
            range_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
            range_end = range_start + timedelta(days=1)
            report_start = today
            report_end = today

        # Sales in range
        sales_qs = Sale.objects.filter(
            created_at__range=(range_start, range_end), status="COMPLETED"
        )
        total_sales = sales_qs.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
        sales_count = sales_qs.count()

        # Profit calculation and day grouping
        sales_with_items = sales_qs.prefetch_related("items__product")
        gross_profit = Decimal("0.00")
        day_groups = defaultdict(list)
        for sale in sales_with_items:
            for item in sale.items.all():
                cost = item.product.cost or Decimal("0")
                gross_profit += (item.unit_price - cost) * item.quantity
            day_groups[sale.created_at.date()].append(sale)

        # Expenses in range
        expenses = (
            Expense.objects.filter(date__gte=report_start, date__lte=report_end).aggregate(
                total=Sum("amount")
            )["total"]
            or Decimal("0.00")
        )

        # Top products
        top_products = (
            SaleItem.objects.filter(sale__in=sales_qs)
            .values("product__name")
            .annotate(total_qty=Sum("quantity"), total_revenue=Sum(F("unit_price") * F("quantity")))
            .order_by("-total_revenue")[:10]
        )

        # Build week_days
        num_days = (report_end - report_start).days + 1
        day_names_es = ["lunes", "martes", "mi\u00e9rcoles", "jueves", "viernes", "s\u00e1bado", "domingo"]

        week_days = []
        for i in range(num_days):
            day = report_start + timedelta(days=i)
            day_sales = day_groups.get(day, [])
            day_total = sum(s.total for s in day_sales) if day_sales else Decimal("0.00")
            day_count = len(day_sales)

            # Top product for this day
            product_qty = defaultdict(int)
            product_revenue = defaultdict(Decimal)
            for sale in day_sales:
                for item in sale.items.all():
                    product_qty[item.product.name] += item.quantity
                    product_revenue[item.product.name] += item.subtotal

            top_product = None
            if product_qty:
                best_name = max(product_qty, key=product_qty.get)
                top_product = {
                    "name": best_name,
                    "units": product_qty[best_name],
                    "revenue": float(product_revenue[best_name]),
                }

            week_days.append({
                "date": day.isoformat(),
                "day_name": day_names_es[day.weekday()],
                "total": float(day_total),
                "count": day_count,
                "top_product": top_product,
            })

        # Summary
        total_week = sum(d["total"] for d in week_days)
        num_days = len(week_days) or 1

        # Compare with previous same-length period
        period_length = (report_end - report_start).days + 1
        prev_start = report_start - timedelta(days=period_length)
        prev_end = report_start
        prev_range_start = timezone.make_aware(datetime.combine(prev_start, datetime.min.time()))
        prev_range_end = timezone.make_aware(datetime.combine(prev_end, datetime.min.time()))
        prev_total = (
            Sale.objects.filter(
                created_at__range=(prev_range_start, prev_range_end), status="COMPLETED"
            ).aggregate(total=Sum("total"))["total"]
            or Decimal("0.00")
        )
        if prev_total > 0:
            change_vs_last_week = round(float(((total_sales - prev_total) / prev_total) * 100), 1)
        elif total_sales > 0:
            change_vs_last_week = 100.0
        else:
            change_vs_last_week = 0.0

        # Fiado pending
        fiado_total = Client.objects.aggregate(total=Sum("current_debt"))["total"] or Decimal("0.00")
        fiado_client_count = Client.objects.filter(current_debt__gt=0).count()

        # Top product overall
        top_product_data = list(top_products[:1])
        top_product = None
        if top_product_data:
            tp = top_product_data[0]
            product_obj = Product.objects.filter(name=tp["product__name"]).first()
            top_product = {
                "name": tp["product__name"],
                "units_sold": tp["total_qty"],
                "revenue": float(tp["total_revenue"]),
                "image": product_obj.image.url if product_obj and product_obj.image else None,
            }

        profit = float(gross_profit - expenses)
        total_sales_float = float(total_sales)
        profit_margin = round((profit / total_sales_float * 100), 1) if total_sales_float > 0 else 0.0
        avg_per_day = round(total_week / num_days, 2)

        return Response({
            "week_days": week_days,
            "summary": {
                "total_week": float(total_sales),
                "change_vs_last_week": change_vs_last_week,
                "avg_per_day": avg_per_day,
            },
            "fiado_pending": {
                "total": float(fiado_total),
                "client_count": fiado_client_count,
            },
            "top_product": top_product,
            "profit": profit,
            "profit_margin": profit_margin,
            "expenses_total": float(expenses),
        })


class RecentActivityView(APIView):
    def get(self, request):
        limit = int(request.query_params.get("limit", 10))

        recent_sales = (
            Sale.objects.filter(status="COMPLETED")
            .order_by("-created_at")
            .select_related("client")[:limit]
        )
        recent_payments = FiadoPayment.objects.all().order_by("-date").select_related("client")[
            :limit
        ]

        activities = []

        for sale in recent_sales:
            activities.append(
                {
                    "id": sale.id,
                    "concept": "Venta",
                    "client_name": sale.client.name if sale.client else "—",
                    "type": "sale",
                    "amount": float(sale.total),
                    "status": "Completado",
                    "date": sale.created_at.strftime("%Y-%m-%d"),
                    "time": sale.created_at.strftime("%H:%M"),
                }
            )

        for payment in recent_payments:
            activities.append(
                {
                    "id": payment.id,
                    "concept": "Pago",
                    "client_name": payment.client.name if payment.client else "—",
                    "type": "payment",
                    "amount": float(payment.amount),
                    "status": "Registrado",
                    "date": payment.date.strftime("%Y-%m-%d"),
                    "time": "12:00",
                }
            )

        activities.sort(key=lambda a: a["date"] + " " + a["time"], reverse=True)

        return Response(activities[:limit])


# ---------------------------------------------------------------------------
# Change Password
# ---------------------------------------------------------------------------


class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        old_password = request.data.get("old_password")
        new_password = request.data.get("new_password")

        if not old_password or not new_password:
            return Response(
                {"detail": "old_password and new_password are required"}, status=400
            )

        if not user.check_password(old_password):
            return Response({"detail": "Contraseña actual incorrecta"}, status=400)

        if len(new_password) < 8:
            return Response({"detail": "La nueva contraseña debe tener al menos 8 caracteres"}, status=400)

        user.set_password(new_password)
        user.save()

        return Response({"detail": "Contraseña actualizada correctamente"})


# ---------------------------------------------------------------------------
# Backup / Restore
# ---------------------------------------------------------------------------


class ExportDbView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Export the database as a gzip-compressed SQLite file."""
        from django.http import FileResponse

        try:
            backup_path = backup_db()
            file_size = os.path.getsize(backup_path)
            response = FileResponse(
                open(backup_path, 'rb'),
                content_type='application/gzip',
                as_attachment=True,
                filename=f"fiadoapp_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db.gz"
            )
            response['Content-Length'] = file_size
            return response
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": f"Error al crear el backup: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ImportDbView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Restore the database from an uploaded backup file."""
        import tempfile

        if 'file' not in request.FILES:
            return Response(
                {"error": "No se envió ningún archivo"},
                status=status.HTTP_400_BAD_REQUEST
            )

        uploaded_file = request.FILES['file']

        if uploaded_file.size > settings.DATA_UPLOAD_MAX_MEMORY_SIZE:
            return Response(
                {"error": f"El archivo excede el límite de "
                          f"{settings.DATA_UPLOAD_MAX_MEMORY_SIZE // 1024 // 1024} MB"},
                status=status.HTTP_400_BAD_REQUEST
            )

        with tempfile.NamedTemporaryFile(suffix='.db.gz', delete=False) as tmp:
            for chunk in uploaded_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            restore_db(tmp_path, create_safety_backup=True)
            return Response({
                "success": True,
                "message": "Base de datos restaurada correctamente"
            })
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": f"Error al restaurar: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


class BackupConfigView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get current backup configuration."""
        config = BackupConfig.get_singleton()
        return Response({
            'enabled': config.enabled,
            'frequency_hours': config.frequency_hours,
            'max_backups': config.max_backups,
            'last_backup': config.last_backup,
            'backup_folder': config.backup_folder or settings.BACKUP_ROOT,
            'db_file_size': get_db_file_size(),
            'db_engine': detect_engine(),
        })

    def put(self, request):
        """Update backup configuration."""
        config = BackupConfig.get_singleton()

        for field in ['enabled', 'frequency_hours', 'max_backups', 'backup_folder']:
            if field in request.data:
                setattr(config, field, request.data[field])

        config.save()
        return Response({'success': True})


# ---------------------------------------------------------------------------
# Exports (XLSX)
# ---------------------------------------------------------------------------


class ExportClientsView(APIView):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        headers = ["ID", "Nombre", "Teléfono", "Email", "Dirección", "Límite Crédito", "Deuda Actual"]
        ws.append(headers)

        for client in Client.objects.all().order_by("name"):
            ws.append(
                [
                    client.id,
                    client.name,
                    client.phone or "",
                    client.email or "",
                    client.address or "",
                    str(client.credit_limit),
                    str(client.current_debt),
                ]
            )

        return _build_xlsx_response(wb, "clientes.xlsx")


class ExportProductsView(APIView):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        headers = [
            "ID",
            "Nombre",
            "Categoría",
            "Precio Venta",
            "Costo",
            "Stock",
            "Stock Mínimo",
            "Código Barras",
        ]
        ws.append(headers)

        for product in Product.objects.all().order_by("name").select_related("category"):
            ws.append(
                [
                    product.id,
                    product.name,
                    product.category.name if product.category else "",
                    str(product.price),
                    str(product.cost),
                    product.stock,
                    product.min_stock,
                    product.barcode or "",
                ]
            )

        return _build_xlsx_response(wb, "productos.xlsx")


class ExportSalesView(APIView):
    def get(self, request):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        headers = [
            "ID Venta",
            "Fecha",
            "Cliente",
            "Total",
            "Método Pago",
            "Estado",
            "Notas",
            "Efectivo Recibido",
            "Vuelto",
        ]
        ws.append(headers)

        qs = Sale.objects.all().order_by("-created_at").select_related("client")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        for sale in qs:
            ws.append(
                [
                    sale.id,
                    sale.created_at.strftime("%Y-%m-%d %H:%M"),
                    sale.client.name if sale.client else "—",
                    str(sale.total),
                    sale.get_payment_method_display(),
                    sale.get_status_display(),
                    sale.notes or "",
                    str(sale.cash_received or ""),
                    str(sale.change_given or ""),
                ]
            )

        return _build_xlsx_response(wb, "ventas.xlsx")
